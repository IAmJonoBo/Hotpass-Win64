"""Enhanced spreadsheet formatting and output generation."""

from __future__ import annotations

from dataclasses import dataclass
from importlib import import_module, util
from pathlib import Path
from typing import Any

import pandas as pd
from pandas.api.types import is_datetime64_any_dtype


def _ensure_pyarrow_parquet_format() -> None:
    """Provide a backwards-compatible ParquetFileFormat attribute for PyArrow."""

    dataset_spec = util.find_spec("pyarrow.dataset")
    if dataset_spec is None:
        return

    dataset_module = import_module("pyarrow.dataset")
    if hasattr(dataset_module, "ParquetFileFormat"):
        return

    parquet_spec = util.find_spec("pyarrow._dataset_parquet")
    if parquet_spec is None:
        return

    parquet_module = import_module("pyarrow._dataset_parquet")
    parquet_format = getattr(parquet_module, "ParquetFileFormat", None)
    if parquet_format is None:
        return

    dataset_module.ParquetFileFormat = parquet_format  # type: ignore[attr-defined]


_ensure_pyarrow_parquet_format()


@dataclass
class OutputFormat:
    """Configuration for output formatting."""

    # Color schemes (as hex colors)
    header_bg_color: str = "366092"
    header_font_color: str = "FFFFFF"
    primary_row_bg: str = "FFFFFF"
    alternate_row_bg: str = "F0F0F0"
    quality_excellent_bg: str = "C6EFCE"
    quality_good_bg: str = "FFEB9C"
    quality_fair_bg: str = "FFC7CE"
    quality_poor_bg: str = "FF0000"

    # Font settings
    font_name: str = "Calibri"
    font_size: int = 11
    header_font_size: int = 12
    header_bold: bool = True

    # Column widths (in Excel units)
    default_column_width: float = 15.0
    auto_size_columns: bool = True
    max_column_width: float = 50.0

    # Other settings
    freeze_header_row: bool = True
    add_filters: bool = True
    zebra_striping: bool = True


def apply_excel_formatting(
    writer: pd.ExcelWriter,
    sheet_name: str,
    df: pd.DataFrame,
    format_config: OutputFormat | None = None,
) -> None:
    """
    Apply formatting to an Excel sheet.

    Args:
        writer: ExcelWriter object
        sheet_name: Name of the sheet to format
        df: DataFrame that was written to the sheet
        format_config: Formatting configuration
    """
    if format_config is None:
        format_config = OutputFormat()

    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Silently skip formatting if openpyxl not available
        return

    worksheet = writer.sheets[sheet_name]

    # Normalise datetime cells to millisecond precision so Excel round-tripping
    # matches the original Pandas values. Pandas writes datetime64 columns with
    # nanosecond precision, but Excel persists datetimes at millisecond
    # precision using half-up rounding. The governance tests expect the
    # workbook to reflect Pandas' own rounding semantics (banker's rounding),
    # so we explicitly coerce values before applying any styling.
    datetime_columns = [column for column in df.columns if is_datetime64_any_dtype(df[column])]
    for column in datetime_columns:
        col_index = df.columns.get_loc(column) + 1
        for row_offset, raw_value in enumerate(df[column], start=2):
            cell = worksheet.cell(row=row_offset, column=col_index)
            if pd.isna(raw_value):
                cell.value = None
                continue

            timestamp = pd.Timestamp(raw_value).round("ms")
            if timestamp.tzinfo is not None:
                timestamp = timestamp.tz_localize(None)

            cell.value = timestamp.to_pydatetime()

    # Format header row
    header_fill = PatternFill(
        start_color=format_config.header_bg_color,
        end_color=format_config.header_bg_color,
        fill_type="solid",
    )
    header_font = Font(
        name=format_config.font_name,
        size=format_config.header_font_size,
        bold=format_config.header_bold,
        color=format_config.header_font_color,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Apply data row formatting
    data_font = Font(name=format_config.font_name, size=format_config.font_size)

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            # Apply zebra striping if enabled
            if format_config.zebra_striping and row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color=format_config.alternate_row_bg,
                    end_color=format_config.alternate_row_bg,
                    fill_type="solid",
                )

    # Apply conditional formatting for quality scores if present
    if "data_quality_score" in df.columns:
        quality_col_idx = df.columns.get_loc("data_quality_score") + 1
        quality_col_letter = get_column_letter(quality_col_idx)

        for row_idx in range(2, len(df) + 2):
            cell = worksheet[f"{quality_col_letter}{row_idx}"]
            try:
                score = float(cell.value) if cell.value is not None else 0
                if score >= 0.8:
                    bg_color = format_config.quality_excellent_bg
                elif score >= 0.6:
                    bg_color = format_config.quality_good_bg
                elif score >= 0.4:
                    bg_color = format_config.quality_fair_bg
                else:
                    bg_color = format_config.quality_poor_bg

                cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            except (ValueError, TypeError):
                pass

    # Adjust column widths
    if format_config.auto_size_columns:
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except (TypeError, AttributeError):
                    pass

            adjusted_width = min(max_length + 2, format_config.max_column_width)
            worksheet.column_dimensions[column_letter].width = max(
                adjusted_width, format_config.default_column_width
            )

    # Freeze header row
    if format_config.freeze_header_row:
        worksheet.freeze_panes = "A2"

    # Add autofilter
    if format_config.add_filters:
        worksheet.auto_filter.ref = worksheet.dimensions


def create_summary_sheet(
    df: pd.DataFrame,
    quality_report: dict[str, Any] | None = None,
) -> pd.DataFrame:
    """
    Create a summary sheet with key metrics and statistics.

    Args:
        df: The main data DataFrame
        quality_report: Optional quality report dictionary

    Returns:
        DataFrame with summary information
    """
    summary_data = []

    # Basic statistics
    summary_data.extend(
        [
            ["Metric", "Value"],
            ["Total Records", len(df)],
            ["", ""],
        ]
    )

    # Quality metrics if available
    if quality_report:
        summary_data.extend(
            [
                ["Quality Metrics", ""],
                ["Invalid Records", quality_report.get("invalid_records", 0)],
                [
                    "Validation Passed",
                    "Yes" if quality_report.get("expectations_passed") else "No",
                ],
                ["", ""],
            ]
        )

    # Field completeness
    summary_data.append(["Field Completeness", "Percentage"])
    for column in df.columns:
        non_null = df[column].notna().sum()
        percentage = (non_null / len(df) * 100) if len(df) > 0 else 0
        summary_data.append([column, f"{percentage:.1f}%"])

    return pd.DataFrame(summary_data)


def export_to_multiple_formats(
    df: pd.DataFrame,
    base_path: Path,
    formats: list[str] | None = None,
    format_config: OutputFormat | None = None,
    quality_report: dict[str, Any] | None = None,
) -> dict[str, Path]:
    """
    Export DataFrame to multiple formats.

    Args:
        df: DataFrame to export
        base_path: Base path for output files (without extension)
        formats: List of formats to export ('excel', 'csv', 'parquet', 'json')
        format_config: Formatting configuration for Excel output
        quality_report: Optional quality report for summary sheet

    Returns:
        Dictionary mapping format names to output file paths
    """
    if formats is None:
        formats = ["excel"]

    output_paths: dict[str, Path] = {}

    for fmt in formats:
        if fmt == "excel":
            output_path = base_path.with_suffix(".xlsx")
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Write main data
                df.to_excel(writer, sheet_name="Data", index=False)
                apply_excel_formatting(writer, "Data", df, format_config)

                # Add summary sheet if quality report available
                if quality_report:
                    summary_df = create_summary_sheet(df, quality_report)
                    summary_df.to_excel(writer, sheet_name="Summary", index=False, header=False)

            output_paths["excel"] = output_path

        elif fmt == "csv":
            output_path = base_path.with_suffix(".csv")
            df.to_csv(output_path, index=False)
            output_paths["csv"] = output_path

        elif fmt == "parquet":
            output_path = base_path.with_suffix(".parquet")
            df.to_parquet(output_path, index=False)
            output_paths["parquet"] = output_path

        elif fmt == "json":
            output_path = base_path.with_suffix(".json")
            df.to_json(output_path, orient="records", indent=2)
            output_paths["json"] = output_path

    return output_paths
