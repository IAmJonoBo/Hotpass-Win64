"""Tests for formatting functionality."""

import pandas as pd
import pytest

pytest.importorskip("frictionless")

from hotpass.formatting import (
    OutputFormat,  # noqa: E402
    apply_excel_formatting,
    create_summary_sheet,
    export_to_multiple_formats,
)


def test_output_format_defaults():
    """Test OutputFormat has sensible defaults."""
    fmt = OutputFormat()

    assert fmt.header_bg_color == "366092"
    assert fmt.font_name == "Calibri"
    assert fmt.font_size == 11
    assert fmt.auto_size_columns is True
    assert fmt.freeze_header_row is True
    assert fmt.zebra_striping is True


def test_output_format_custom_values():
    """Test OutputFormat with custom values."""
    fmt = OutputFormat(
        header_bg_color="FF0000",
        font_name="Arial",
        font_size=14,
        auto_size_columns=False,
    )

    assert fmt.header_bg_color == "FF0000"
    assert fmt.font_name == "Arial"
    assert fmt.font_size == 14
    assert fmt.auto_size_columns is False


def test_create_summary_sheet_basic():
    """Test creating a summary sheet with basic data."""
    df = pd.DataFrame(
        {
            "name": ["Alice", "Bob", None],
            "email": ["alice@test.com", "bob@test.com", "charlie@test.com"],
            "phone": ["123", None, "456"],
        }
    )

    summary_df = create_summary_sheet(df)

    assert len(summary_df) > 0
    # Should contain total records
    assert any("Total Records" in str(row) for _, row in summary_df.iterrows())


def test_create_summary_sheet_with_quality_report():
    """Test creating summary sheet with quality report data."""
    df = pd.DataFrame({"col1": [1, 2, 3], "col2": [4, 5, 6]})

    quality_report = {
        "invalid_records": 2,
        "expectations_passed": True,
    }

    summary_df = create_summary_sheet(df, quality_report)

    assert len(summary_df) > 0
    # Should contain quality metrics
    assert any("Invalid Records" in str(row) for _, row in summary_df.iterrows())
    assert any("Yes" in str(row) for _, row in summary_df.iterrows())


def test_create_summary_sheet_completeness_percentage():
    """Test that summary sheet calculates field completeness."""
    df = pd.DataFrame(
        {
            "col1": [1, 2, None, 4, 5],  # 80% complete
            "col2": [1, 2, 3, 4, 5],  # 100% complete
        }
    )

    summary_df = create_summary_sheet(df)

    # Convert to string for easier searching
    summary_text = summary_df.to_string()

    assert "col1" in summary_text
    assert "col2" in summary_text
    # Should have percentage values
    assert "%" in summary_text


def test_export_to_excel(tmp_path):
    """Test exporting to Excel format."""
    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(df, base_path, formats=["excel"])

    assert "excel" in result
    assert result["excel"].exists()
    assert result["excel"].suffix == ".xlsx"


def test_export_to_csv(tmp_path):
    """Test exporting to CSV format."""
    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(df, base_path, formats=["csv"])

    assert "csv" in result
    assert result["csv"].exists()
    assert result["csv"].suffix == ".csv"

    # Verify contents
    loaded = pd.read_csv(result["csv"])
    assert len(loaded) == 2
    assert list(loaded.columns) == ["name", "age"]


def test_export_to_parquet(tmp_path):
    """Test exporting to Parquet format."""
    pytest.importorskip("pyarrow", reason="pyarrow not installed")

    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(df, base_path, formats=["parquet"])

    assert "parquet" in result
    assert result["parquet"].exists()
    assert result["parquet"].suffix == ".parquet"

    # Verify contents
    loaded = pd.read_parquet(result["parquet"])
    assert len(loaded) == 2


def test_export_to_json(tmp_path):
    """Test exporting to JSON format."""
    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(df, base_path, formats=["json"])

    assert "json" in result
    assert result["json"].exists()
    assert result["json"].suffix == ".json"

    # Verify contents
    loaded = pd.read_json(result["json"])
    assert len(loaded) == 2


def test_export_to_multiple_formats_all(tmp_path):
    """Test exporting to all formats at once."""
    pytest.importorskip("pyarrow", reason="pyarrow not installed")

    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(df, base_path, formats=["excel", "csv", "parquet", "json"])

    assert len(result) == 4
    assert all(path.exists() for path in result.values())


def test_export_with_quality_report(tmp_path):
    """Test exporting Excel with quality report creates summary sheet."""
    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    quality_report = {
        "invalid_records": 0,
        "expectations_passed": True,
    }

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(
        df, base_path, formats=["excel"], quality_report=quality_report
    )

    assert result["excel"].exists()

    # Verify summary sheet was created
    with pd.ExcelFile(result["excel"]) as xls:
        assert "Summary" in xls.sheet_names
        assert "Data" in xls.sheet_names


def test_export_defaults_to_excel(tmp_path):
    """Test that export defaults to Excel format when no formats specified."""
    df = pd.DataFrame({"name": ["Alice"], "age": [25]})

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(df, base_path)

    assert "excel" in result
    assert result["excel"].exists()


def test_apply_excel_formatting_with_quality_score(tmp_path):
    """Test applying formatting with quality score column."""
    df = pd.DataFrame(
        {
            "name": ["Alice", "Bob", "Charlie"],
            "data_quality_score": [0.9, 0.5, 0.2],  # Excellent, fair, poor
        }
    )

    output_path = tmp_path / "test.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        apply_excel_formatting(writer, "Data", df)

    assert output_path.exists()

    # Verify formatting was applied
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    ws = wb["Data"]

    # Check that header row exists
    assert ws["A1"].value == "name"
    assert ws["B1"].value == "data_quality_score"

    # Check data exists
    assert ws["A2"].value == "Alice"


def test_apply_excel_formatting_without_quality_score(tmp_path):
    """Test applying formatting without quality score column."""
    df = pd.DataFrame({"name": ["Alice", "Bob"], "age": [25, 30]})

    output_path = tmp_path / "test.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        apply_excel_formatting(writer, "Data", df)

    assert output_path.exists()


def test_apply_excel_formatting_custom_config(tmp_path):
    """Test applying custom formatting configuration."""
    df = pd.DataFrame({"name": ["Alice"], "age": [25]})

    custom_format = OutputFormat(
        header_bg_color="FF0000",
        font_size=14,
        zebra_striping=False,
        freeze_header_row=False,
    )

    output_path = tmp_path / "test.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        apply_excel_formatting(writer, "Data", df, custom_format)

    assert output_path.exists()


def test_apply_excel_formatting_with_invalid_quality_scores(tmp_path):
    """Test formatting handles invalid quality score values gracefully."""
    df = pd.DataFrame(
        {
            "name": ["Alice", "Bob", "Charlie"],
            "data_quality_score": [0.9, None, "invalid"],  # Mix of valid/invalid
        }
    )

    output_path = tmp_path / "test.xlsx"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        apply_excel_formatting(writer, "Data", df)

    assert output_path.exists()


def test_create_summary_sheet_empty_dataframe():
    """Test creating summary sheet with empty DataFrame."""
    df = pd.DataFrame()

    summary_df = create_summary_sheet(df)

    assert len(summary_df) > 0
    # Should still have headers/structure
    assert any("Total Records" in str(row) for _, row in summary_df.iterrows())


def test_export_with_custom_format_config(tmp_path):
    """Test export with custom formatting configuration."""
    df = pd.DataFrame({"name": ["Alice"], "age": [25]})

    custom_format = OutputFormat(font_name="Arial", font_size=16)

    base_path = tmp_path / "output"
    result = export_to_multiple_formats(
        df, base_path, formats=["excel"], format_config=custom_format
    )

    assert result["excel"].exists()
