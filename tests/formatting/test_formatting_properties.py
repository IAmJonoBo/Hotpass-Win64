"""Property-based tests for formatting stability and encoding coverage."""

from __future__ import annotations

import io
import string
from collections.abc import Callable
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd
import pytest
from hotpass.formatting import apply_excel_formatting

from tests.helpers.hypothesis import DrawFn, SearchStrategy, composite, given, settings, st


def expect(condition: bool, message: str) -> None:
    if not condition:
        pytest.fail(message)


_EXCEL_SAFE_TEXT = st.text(
    alphabet=st.characters(
        blacklist_categories=("Cs",),
        blacklist_characters={chr(codepoint) for codepoint in range(32)} | {chr(127)},
    ),
    max_size=12,
)


@settings(max_examples=10, deadline=None)
@given(st.lists(_EXCEL_SAFE_TEXT, min_size=1, max_size=4))
def test_apply_excel_formatting_preserves_unicode(values: list[str]) -> None:
    frame = pd.DataFrame({"text": values})
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Data", index=False)
        apply_excel_formatting(writer, "Data", frame)

    workbook = openpyxl.load_workbook(io.BytesIO(buffer.getvalue()))
    sheet = workbook["Data"]
    observed_column = next(
        sheet.iter_cols(min_row=2, max_row=len(values) + 1, min_col=1, max_col=1)
    )
    observed = [cell.value if cell.value is not None else "" for cell in observed_column]
    expect(observed == values, "Unicode cell values should survive formatting round-trip")


@composite
def _formatting_frames(draw: DrawFn) -> pd.DataFrame:
    row_count = draw(st.integers(min_value=1, max_value=4))
    names = draw(
        st.lists(
            st.text(alphabet=string.ascii_letters + " ", min_size=1, max_size=20),
            min_size=row_count,
            max_size=row_count,
        )
    )
    data: dict[str, Any] = {"name": names}

    if draw(st.booleans()):
        scores = draw(
            st.lists(
                st.floats(min_value=0.0, max_value=1.0, allow_nan=False, allow_infinity=False),
                min_size=row_count,
                max_size=row_count,
            )
        )
        data["data_quality_score"] = scores

    if draw(st.booleans()):
        timestamps = draw(
            st.lists(
                st.datetimes(min_value=datetime(2018, 1, 1), max_value=datetime(2035, 12, 31)),
                min_size=row_count,
                max_size=row_count,
            )
        )
        data["observed_at"] = timestamps

    return pd.DataFrame(data)


_FORMATTING_FRAMES_FACTORY: Callable[[], SearchStrategy[pd.DataFrame]] = _formatting_frames
_FORMATTING_FRAMES: SearchStrategy[pd.DataFrame] = _FORMATTING_FRAMES_FACTORY()


@settings(max_examples=12, deadline=None)
@given(_FORMATTING_FRAMES)
def test_apply_excel_formatting_handles_optional_columns(frame: pd.DataFrame) -> None:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name="Data", index=False)
        apply_excel_formatting(writer, "Data", frame)

    workbook = openpyxl.load_workbook(io.BytesIO(buffer.getvalue()))
    sheet = workbook["Data"]

    expect(
        sheet.max_row == len(frame) + 1,
        "Row count should be preserved after formatting",
    )

    if "observed_at" in frame.columns:
        col_index = frame.columns.get_loc("observed_at") + 1
        observed_cells = [
            pd.Timestamp(cell.value).to_pydatetime() if cell.value is not None else None
            for cell in next(
                sheet.iter_cols(
                    min_row=2,
                    max_row=len(frame) + 1,
                    min_col=col_index,
                    max_col=col_index,
                )
            )
        ]
        expected = [
            pd.Timestamp(value).round("ms").to_pydatetime() if pd.notna(value) else None
            for value in frame["observed_at"]
        ]
        expect(
            observed_cells == expected,
            "Datetime cells must remain intact after formatting",
        )
